"""Excel report renderer using openpyxl.

Produces a workbook with three sheets:
  Summary  — KPI table (peak TPS, avg TPS, latency, error rate)
  TimeSeries — per-observation raw data (analysts copy-paste to their models)
  Config   — deployment metadata (URL, environment, period)
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import ReportData

_WHITE = "FFFFFFFF"


def _hex_to_argb(hex_colour: str) -> str:
    """Convert #rrggbb or rrggbb → FFRRGGBB (openpyxl ARGB format)."""
    h = hex_colour.lstrip("#")
    return "FF" + h.upper()


def _header_font(bold: bool = True) -> Font:
    return Font(bold=bold, color=_WHITE, name="Arial")


def _header_fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=_hex_to_argb(hex_colour))


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def render_excel(data: ReportData, primary: str, secondary: str, company: str) -> bytes:
    """Return Excel workbook bytes."""
    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    title_font = Font(bold=True, size=14, name="Arial", color=_hex_to_argb(primary))
    ws_sum["A1"] = f"{company} — Stub Performance Report"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = f"Project: {data.project_name} | Stub: {data.stub_name}"
    ws_sum["A2"].font = Font(italic=True, name="Arial", size=10)
    ws_sum["A3"] = f"Generated: {data.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
    ws_sum["A3"].font = Font(name="Arial", size=10)

    headers = ["Metric", "Value"]
    kpis = [
        ("Peak TPS", f"{data.peak_tps:,.1f}"),
        ("Average TPS", f"{data.avg_tps:,.1f}"),
        ("Avg Latency (ms)", f"{data.avg_latency_ms:.2f}"),
        ("p95 Latency (ms)", f"{data.p95_latency_ms:.2f}"),
        ("Total Requests", f"{data.total_requests:,}"),
        ("Error Rate (%)", f"{data.error_rate_pct:.4f}"),
        ("Environment", data.environment),
        ("Stub URL", data.stub_url),
        ("Report Period (h)", str(data.report_period_hours)),
    ]

    header_fill = _header_fill(primary)
    for col, h in enumerate(headers, start=1):
        cell = ws_sum.cell(row=5, column=col, value=h)
        cell.font = _header_font()
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    alt_fill = _header_fill(secondary)
    for row_idx, (label, value) in enumerate(kpis, start=6):
        ws_sum.cell(row=row_idx, column=1, value=label).font = Font(bold=True, name="Arial", size=11)
        ws_sum.cell(row=row_idx, column=2, value=value).alignment = Alignment(horizontal="right")
        if row_idx % 2 == 0:
            for col in (1, 2):
                ws_sum.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFE8F4FF")

    _set_col_widths(ws_sum, [30, 30])

    # ── TimeSeries sheet ──────────────────────────────────────────────────────
    ws_ts = wb.create_sheet("TimeSeries")
    ts_headers = ["Time (UTC)", "TPS", "Avg Latency (ms)", "Error Rate (%)"]
    for col, h in enumerate(ts_headers, start=1):
        cell = ws_ts.cell(row=1, column=col, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill(secondary)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, point in enumerate(data.points, start=2):
        ws_ts.cell(row=row_idx, column=1, value=point.time.strftime("%Y-%m-%d %H:%M"))
        ws_ts.cell(row=row_idx, column=2, value=round(point.tps, 2))
        ws_ts.cell(row=row_idx, column=3, value=round(point.latency_avg_ms, 3))
        ws_ts.cell(row=row_idx, column=4, value=round(point.error_rate * 100, 4))

    _set_col_widths(ws_ts, [22, 14, 20, 18])

    # ── Config sheet ──────────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("Config")
    cfg_rows = [
        ("deployment_id", data.deployment_id),
        ("project_id", data.project_id),
        ("stub_id", data.stub_id),
        ("stub_url", data.stub_url),
        ("environment", data.environment),
        ("report_period_hours", data.report_period_hours),
        ("generated_at", data.generated_at.isoformat()),
    ]
    for row_idx, (k, v) in enumerate(cfg_rows, start=1):
        ws_cfg.cell(row=row_idx, column=1, value=k).font = Font(bold=True, name="Arial")
        ws_cfg.cell(row=row_idx, column=2, value=str(v))
    _set_col_widths(ws_cfg, [30, 50])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
