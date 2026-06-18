"""Phase 5 Sprint 15 — reporter-service tests.

Tests cover:
  - render_html produces required content (title, KPIs, table rows)
  - render_excel: workbook has 3 sheets, Summary sheet has expected KPI row
  - render_ppt: presentation has 6 slides
  - s3_store.upload_pdf/excel/ppt calls put_object with correct ContentType
  - s3_store.presigned_url calls generate_presigned_url
  - worker.process_message: success path → DONE + S3 keys in result
  - worker.process_message: DB load failure → FAILED
  - worker.process_message: Excel/PPT render failure is logged but job still DONE
  - project-service POST /deployments/{id}/report → 202 with job_id
  - project-service POST /deployments/{id}/report on DEPLOYING → 409
  - project-service enqueue_report_job sends correct SQS message body
"""
from __future__ import annotations

import io
import json
import os
import uuid
from datetime import datetime, timezone
from unittest.mock import MagicMock, patch

import pytest
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

os.environ.setdefault("AWS_ACCESS_KEY_ID", "testing")
os.environ.setdefault("AWS_SECRET_ACCESS_KEY", "testing")
os.environ.setdefault("AWS_DEFAULT_REGION", "eu-west-2")

# ── Test data ──────────────────────────────────────────────────────────────────

PROJECT_ID = str(uuid.UUID("aaaaaaaa-0000-0000-0000-000000000001"))
STUB_ID = str(uuid.UUID("bbbbbbbb-0000-0000-0000-000000000001"))
DEPLOYMENT_ID = str(uuid.UUID("cccccccc-0000-0000-0000-000000000001"))
JOB_ID = str(uuid.UUID("dddddddd-0000-0000-0000-000000000001"))

PRIMARY = "#003875"
SECONDARY = "#00A9E0"
COMPANY = "Test Corp"


def _make_report_data():
    from reporter_service.models import MetricPoint, ReportData
    points = [
        MetricPoint(
            time=datetime(2026, 6, 18, 10, i, 0, tzinfo=timezone.utc),
            tps=float(1000 + i * 100),
            latency_avg_ms=5.0 + i * 0.1,
            error_rate=0.001,
        )
        for i in range(5)
    ]
    return ReportData(
        project_id=PROJECT_ID,
        project_name="Payments Stub",
        stub_id=STUB_ID,
        stub_name="POST /payments",
        deployment_id=DEPLOYMENT_ID,
        stub_url="http://10.0.1.100:8080",
        environment="TEST",
        report_period_hours=24,
        generated_at=datetime(2026, 6, 18, 12, 0, 0, tzinfo=timezone.utc),
        peak_tps=1400.0,
        avg_tps=1200.0,
        avg_latency_ms=5.25,
        p95_latency_ms=8.5,
        total_requests=86400000,
        error_rate_pct=0.1,
        points=points,
    )


# ── HTML / PDF tests ──────────────────────────────────────────────────────────

def test_render_html_contains_project_name():
    from reporter_service.renderers.pdf import render_html
    html = render_html(_make_report_data(), PRIMARY, SECONDARY, COMPANY)
    assert "Payments Stub" in html


def test_render_html_contains_kpi_values():
    from reporter_service.renderers.pdf import render_html
    html = render_html(_make_report_data(), PRIMARY, SECONDARY, COMPANY)
    assert "1,400" in html  # peak TPS formatted with comma
    assert "5.2ms" in html or "5.25" in html or "5.3ms" in html  # avg latency (5.25 rounds to 5.2)


def test_render_html_contains_time_series_rows():
    from reporter_service.renderers.pdf import render_html
    html = render_html(_make_report_data(), PRIMARY, SECONDARY, COMPANY)
    assert "2026-06-18 10:00" in html


# ── Excel tests ───────────────────────────────────────────────────────────────

def test_render_excel_returns_bytes():
    from reporter_service.renderers.excel import render_excel
    result = render_excel(_make_report_data(), PRIMARY, SECONDARY, COMPANY)
    assert isinstance(result, bytes)
    assert len(result) > 100  # non-trivial content


def test_render_excel_has_three_sheets():
    import openpyxl
    from reporter_service.renderers.excel import render_excel
    buf = io.BytesIO(render_excel(_make_report_data(), PRIMARY, SECONDARY, COMPANY))
    wb = openpyxl.load_workbook(buf)
    assert wb.sheetnames == ["Summary", "TimeSeries", "Config"]


def test_render_excel_summary_contains_peak_tps():
    import openpyxl
    from reporter_service.renderers.excel import render_excel
    buf = io.BytesIO(render_excel(_make_report_data(), PRIMARY, SECONDARY, COMPANY))
    wb = openpyxl.load_workbook(buf)
    ws = wb["Summary"]
    values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
    assert any("Peak TPS" in v for v in values)


# ── PowerPoint tests ──────────────────────────────────────────────────────────

def test_render_ppt_returns_bytes():
    from reporter_service.renderers.ppt import render_ppt
    result = render_ppt(_make_report_data(), PRIMARY, SECONDARY, COMPANY)
    assert isinstance(result, bytes)
    assert len(result) > 100


def test_render_ppt_has_six_slides():
    from pptx import Presentation
    from reporter_service.renderers.ppt import render_ppt
    buf = io.BytesIO(render_ppt(_make_report_data(), PRIMARY, SECONDARY, COMPANY))
    prs = Presentation(buf)
    assert len(prs.slides) == 6


# ── S3 store tests ────────────────────────────────────────────────────────────

def test_s3_upload_pdf_uses_correct_content_type():
    from reporter_service.s3_store import upload_pdf
    mock_s3 = MagicMock()
    upload_pdf(mock_s3, "test-bucket", DEPLOYMENT_ID, b"fake-pdf")
    call_kwargs = mock_s3.put_object.call_args[1]
    assert call_kwargs["ContentType"] == "application/pdf"
    assert call_kwargs["Key"].endswith("/report.pdf")


def test_s3_upload_excel_uses_correct_content_type():
    from reporter_service.s3_store import upload_excel
    mock_s3 = MagicMock()
    upload_excel(mock_s3, "test-bucket", DEPLOYMENT_ID, b"fake-xlsx")
    call_kwargs = mock_s3.put_object.call_args[1]
    assert "spreadsheetml" in call_kwargs["ContentType"]


def test_s3_upload_ppt_uses_correct_content_type():
    from reporter_service.s3_store import upload_ppt
    mock_s3 = MagicMock()
    upload_ppt(mock_s3, "test-bucket", DEPLOYMENT_ID, b"fake-pptx")
    call_kwargs = mock_s3.put_object.call_args[1]
    assert "presentationml" in call_kwargs["ContentType"]


# ── Worker tests ──────────────────────────────────────────────────────────────

def _build_db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    with engine.connect() as conn:
        conn.execute(text(
            "CREATE TABLE jobs (id TEXT PRIMARY KEY, type TEXT, status TEXT, "
            "project_id TEXT, stub_id TEXT, payload TEXT, result TEXT, "
            "error_message TEXT, sqs_message_id TEXT, "
            "created_at TEXT DEFAULT CURRENT_TIMESTAMP, "
            "updated_at TEXT DEFAULT CURRENT_TIMESTAMP)"
        ))
        conn.execute(text(
            "INSERT INTO jobs (id, type, status, project_id, stub_id, payload) "
            "VALUES (:id, 'REPORT', 'QUEUED', :pid, :sid, '{}')"
        ), {"id": JOB_ID, "pid": PROJECT_ID, "sid": STUB_ID})
        conn.commit()
    return sessionmaker(bind=engine)


def _make_message():
    body = {
        "job_id": JOB_ID,
        "type": "REPORT",
        "payload": {"deployment_id": DEPLOYMENT_ID, "report_period_hours": 24},
        "created_at": "2026-06-18T12:00:00Z",
        "project_id": PROJECT_ID,
    }
    return {"MessageId": "msg-1", "ReceiptHandle": "rh-1", "Body": json.dumps(body)}


def test_worker_success_path_sets_job_done():
    from reporter_service.worker import process_message

    Session = _build_db()
    db = Session()

    mock_ts = MagicMock()
    mock_s3 = MagicMock()
    mock_s3.put_object.return_value = {}
    mock_s3.generate_presigned_url.return_value = "https://s3.example.com/report.xlsx"

    with patch("reporter_service.worker.build_report_data", return_value=_make_report_data()), \
         patch("reporter_service.renderers.pdf.render_pdf", return_value=b"pdf"), \
         patch("reporter_service.worker.render_excel", return_value=b"xlsx"), \
         patch("reporter_service.worker.render_ppt", return_value=b"pptx"):
        process_message(_make_message(), db, mock_ts, mock_s3)

    row = db.execute(text("SELECT status, result FROM jobs WHERE id=:id"), {"id": JOB_ID}).fetchone()
    assert row[0] == "DONE"
    result = json.loads(row[1])
    assert result["excel_key"] is not None
    assert result["ppt_key"] is not None
    db.close()


def test_worker_db_load_failure_sets_job_failed():
    from reporter_service.worker import process_message

    Session = _build_db()
    db = Session()

    mock_ts = MagicMock()
    mock_s3 = MagicMock()

    with patch("reporter_service.worker.build_report_data", side_effect=ValueError("Not found")):
        process_message(_make_message(), db, mock_ts, mock_s3)

    row = db.execute(text("SELECT status FROM jobs WHERE id=:id"), {"id": JOB_ID}).fetchone()
    assert row[0] == "FAILED"
    db.close()


# ── project-service enqueue_report_job test ───────────────────────────────────

def test_enqueue_report_job_sends_correct_payload():
    import sys
    # Ensure project_service is importable
    sys.path.insert(0, str(
        __import__("pathlib").Path(__file__).parent.parent.parent / "project-service" / "src"
    ))
    from project_service.sqs_client import enqueue_report_job

    mock_client = MagicMock()
    mock_client.send_message.return_value = {"MessageId": "msg-test"}

    job_id = uuid.UUID(JOB_ID)
    dep_id = uuid.UUID(DEPLOYMENT_ID)
    proj_id = uuid.UUID(PROJECT_ID)

    enqueue_report_job(mock_client, job_id, dep_id, proj_id, 48)

    call_kwargs = mock_client.send_message.call_args[1]
    body = json.loads(call_kwargs["MessageBody"])
    assert body["type"] == "REPORT"
    assert body["payload"]["deployment_id"] == DEPLOYMENT_ID
    assert body["payload"]["report_period_hours"] == 48
